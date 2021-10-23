#!/usr/bin/env python

import sys
import argparse
import string
import re
import numpy as np
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.styles import PatternFill

parser = argparse.ArgumentParser(description='Analyze Map file of Kail-MDk')
parser.add_argument("-o", "--Output", type=str, help="output directory")
parser.add_argument("-i", "--Input", type=str, help="input map file")

args = parser.parse_args()

if not args.Input or not args.Output:
    if not args.Input:
        print('No input file ...')
    if not args.Output:
        print('No output directory ...')

    sys.exit(1)

symbol_table = []
region_list = []

def parse_func_symbols():
    global  symbol_table
    global  region_list
    is_start = False
    fout = 0
    exec_region_cnt = 0
    with open(args.Input, 'r') as fin:
        for line in fin:
            symbol_item = []

            if is_start == False:
                # skip dummp context
                if "Memory Map of the image" in line:
                    is_start = True
                    continue

                continue

            if 'Execution Region' in line:
                if fout:
                    fout.close()

                matchObj = re.search(r'^\s+ Execution Region (\w+) \(Exec base: 0x([0-9a-fA-F]+), Load base: 0x([0-9a-fA-F]+), Size: 0x([0-9a-fA-F]+), Max: 0x([0-9a-fA-F]+)', line)
                if matchObj:
                    region_name = '_' + matchObj.group(1) + '_'

                    a_region = []
                    a_region.append(int(matchObj.group(2), 16)) # base address
                    a_region.append(int(matchObj.group(5), 16)) # region max size
                    a_region.append(matchObj.group(1))          # region name
                    region_list.append(a_region)
                    print("%s: base= 0x%08X, size= %d, max= 0x%08X" % (matchObj.group(1), int(matchObj.group(2), 16), int(matchObj.group(4), 16), int(matchObj.group(5), 16)))

                    # open output file
                    fout = open(args.Output + region_name + '.txt', 'w')
                    fout.write("This is line %d\n" % exec_region_cnt)
                    exec_region_cnt = exec_region_cnt + 1

            if 'PAD' in line:
                continue

            matchObj = re.search(r'^\s+ 0x([0-9a-fA-F]+)\s+ 0x([0-9a-fA-F]+)\s+ 0x([0-9a-fA-F]+)\s+ ([a-zA-Z]+)\s+ ([RO]+)\s+ \d+\s+(.+)\.o\)$', line)
            if matchObj:
                symbol_item.append(int(matchObj.group(1), 16))  # base address
                symbol_item.append(int(matchObj.group(3), 16))  # symbol size
                symbol_item.append(matchObj.group(5))           # type, RO
                symbol_item.append(matchObj.group(6) + '.o)')   # object name
                symbol_table.append(symbol_item)
                # print("addr= 0x%08X, size= %5d, %s, %s.o)" %(int(matchObj.group(1), 16), int(matchObj.group(3), 16), matchObj.group(5), matchObj.group(6)))

            matchObj = re.search(r'^\s+ 0x([0-9a-fA-F]+)\s+ 0x([0-9a-fA-F]+)\s+ 0x([0-9a-fA-F]+)\s+ ([a-zA-Z]+)\s+ ([RO]+)\s+ \d+\s+(.+)\.o$', line)
            if matchObj:
                symbol_item.append(int(matchObj.group(1), 16))  # base address
                symbol_item.append(int(matchObj.group(3), 16))  # symbol size
                symbol_item.append(matchObj.group(5))           # type, RO
                symbol_item.append(matchObj.group(6) + '.o')    # object name
                symbol_table.append(symbol_item)
                # print("addr= 0x%08X, size= %5d, %s, %s.o" %(int(matchObj.group(1), 16), int(matchObj.group(3), 16), matchObj.group(5), matchObj.group(6)))


            if '===========================================' in line:
                break

def parse_data_symbols():
    global  symbol_table
    is_start = False
    with open(args.Input, 'r') as fin:
        for line in fin:
            symbol_item = []

            if is_start == False:
                # skip dummp context
                if "Image Symbol Table" in line:
                    is_start = True
                    continue

                continue

            matchObj = re.search(r'^\s+ ([\w.]+)\s+ 0x([0-9a-fA-F]+)\s+ ([Data]+)\s+ (\d+) (.+)$', line)
            if matchObj:
                symbol_item.append(int(matchObj.group(2), 16))  # base address
                symbol_item.append(int(matchObj.group(4), 16))  # symbol size
                symbol_item.append(matchObj.group(3))           # type, Data
                symbol_item.append(matchObj.group(1))           # object name
                symbol_table.append(symbol_item)
                # print("addr= 0x%08X, size= %5d, %s, %s %s" %(int(matchObj.group(2), 16), int(matchObj.group(4), 10), matchObj.group(3), matchObj.group(1), matchObj.group(5)))

            if 'STACK' in line:
                matchObj = re.search(r'^\s+ ([\w.]+)\s+ 0x([0-9a-fA-F]+)\s+ ([Section]+)\s+ (\d+) (.+)$', line)
                if matchObj:
                    symbol_item.append(int(matchObj.group(2), 16))  # base address
                    symbol_item.append(int(matchObj.group(4), 16))  # symbol size
                    symbol_item.append(matchObj.group(3))           # type, Data
                    symbol_item.append(matchObj.group(1))           # object name
                    symbol_table.append(symbol_item)
                    # print("addr= 0x%08X, size= %5d, Data, %s %s" %(int(matchObj.group(2), 16), int(matchObj.group(4), 10), matchObj.group(1), matchObj.group(5)))


            if '===========================================' in line:
                break


def save_csv_file(fout, sym_start_addr, sym_size, type, sym_name):

        fout.write("%s, %6d, %s, %s\n" % ("{0:#0{1}x}".format(sym_start_addr, 8), sym_size, type, sym_name))


def draw(symbol_table, addr_start, addr_end):
    # colors = ['#3B9DD3', '#41ADE8', '#48BEFF', '#44D5FF', '#40EBFF', '#40E0CF', '#43C59E', '#42B091', '#409B83', '#51A48E']
    colors = ['r', 'pink', 'orange', 'y', 'g', 'b', 'deeppink', 'purple', 'brown', 'black']

    fig, gnt = plt.subplots()

    # gnt.set_xlim(addr_start, addr_end)
    # xlabels = map(lambda t: '0x%08X' % int(t), gnt.get_xticks())
    # gnt.set_xticklabels(xlabels);
    # plt.xticks(rotation = 45)

    target_dpi = 200
    max_addr = 0
    sym_cnt = 1
    for j in range(len(symbol_table)):
        sym_base_addr = symbol_table[j][0]
        sym_size      = symbol_table[j][1]
        sym_end_addr  = sym_base_addr + sym_size
        sym_name      = symbol_table[j][3]
        sym_type      = symbol_table[j][2]
        if (addr_start <= sym_base_addr and sym_end_addr <= addr_end) or \
           (sym_base_addr <= addr_end and sym_end_addr >= addr_end):

            if sym_end_addr > max_addr:
                max_addr = sym_end_addr

            sym_color = "green"
            if "RO" in sym_type:
                sym_color = "orange"

            sym_name = sym_name.replace('$', '#')

            if sym_size == 0:
                plt.barh(sym_cnt, 2, left=sym_base_addr, height=1.0, color="red")
                plt.text(sym_base_addr + 2, sym_cnt, '%s @0x%08X' % (sym_name, sym_base_addr), color="red", size=6)

            else:
                plt.barh(sym_cnt, sym_size, left=sym_base_addr, height=1.0, color=sym_color)
                plt.text(sym_base_addr + sym_size, sym_cnt, '%s @0x%08X~0x%08X' % (sym_name, sym_base_addr, sym_base_addr + sym_size), color="black", size=6)

            sym_cnt = sym_cnt + 1


    gnt.set_xlim(addr_start, max_addr)
    # gnt.set_xlim(addr_start, addr_end)
    xlabels = map(lambda t: '0x%08X' % int(t), gnt.get_xticks())
    gnt.set_xticklabels(xlabels);
    plt.xticks(rotation = 45)
    # gnt.xaxis.set_ticks(np.arange(addr_start, max_addr, 64))

    # plt.xlabel('address')
    # plt.ylabel('symbols')
    gnt.set_xlabel('address (used: %d KB, free: %d KB)' % ((max_addr - addr_start) / 1024, (addr_end - max_addr) / 1024))
    gnt.set_ylabel('symbols')

    plt.grid(True)
    plt.tight_layout()

    # fig.set_figheight(6) # 600
    # fig.set_figwidth(8)  # 800

    fig.set_figheight(10.80) # 1080
    fig.set_figwidth(19.20)  # 1920

    out_path = args.Output + '/' + args.Input
    out_path = out_path.replace('.map', '.png')
    pos = out_path.find('.png')
    out_path = out_path[:pos] + "%08X_%08X" % (addr_start, addr_end) + out_path[pos:]
    plt.savefig(out_path, dpi=target_dpi)
    plt.show()


def main():
    global  symbol_table
    global  region_list
    parse_func_symbols()
    parse_data_symbols()

    region_list = np.array(region_list, dtype=object)
    region_list_sort = region_list[region_list[:, 0].argsort()] # sort by address

    symbol_table = np.array(symbol_table, dtype=object)
    symbol_table_sort = symbol_table[symbol_table[:, 0].argsort()] # sort by address

    # # dump elements
    # for i in range(len(symbol_table_sort)):
    #     print(symbol_table_sort[i])
    #
    # for i in range(len(region_list_sort)):
    #     print(region_list_sort[i])

    wbook = Workbook()

    for i in range(len(region_list_sort)):
        addr_start = region_list_sort[i][0]
        addr_end   = region_list_sort[i][0] + region_list_sort[i][1]
        # print("0x%08x ~ 0x%08X" % (addr_star, addr_end))

        draw(symbol_table_sort, addr_start, addr_end)

        sheet_name = "%08X_%08X" % (addr_start, addr_end)
        wsheet = wbook.create_sheet(sheet_name, 0)
        row = ["address", "size", "type", "obj_name"]
        wsheet.append(row)

        for j in range(len(symbol_table_sort)):
            sym_base_addr = symbol_table_sort[j][0]
            sym_size      = symbol_table_sort[j][1]
            sym_end_addr  = sym_base_addr + sym_size

            if (addr_start <= sym_base_addr and sym_end_addr <= addr_end) or \
               (sym_base_addr <= addr_end and sym_end_addr >= addr_end):
                wsheet.append(["{0:#0{1}x}".format(sym_base_addr, 8), sym_size, symbol_table_sort[j][2], symbol_table_sort[j][3]])



        # out_path = "%s/%08X_%08X.csv" % (args.Output, addr_star, addr_end)
        # with open(out_path, 'w') as fout:
        #     fout.write("address, size, type, obj_name\n")
        #
        #     for j in range(len(symbol_table_sort)):
        #         sym_base_addr = symbol_table_sort[j][0]
        #         sym_size      = symbol_table_sort[j][1]
        #         sym_end_addr  = sym_base_addr + sym_size
        #
        #         if (addr_star <= sym_base_addr and sym_end_addr <= addr_end) or \
        #            (sym_base_addr <= addr_end and sym_end_addr >= addr_end):
        #             save_csv_file(fout, sym_base_addr, sym_size, symbol_table_sort[j][2], symbol_table_sort[j][3])


    out_path = args.Output + '/' + args.Input
    out_path = out_path.replace('.map', '.xlsx')
    wbook.save(out_path)


if __name__ == "__main__":
    main()